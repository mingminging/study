from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import requests
from bs4 import BeautifulSoup
import time
from openpyxl import load_workbook

URLN="https://naver.com"
URLD="https://daum.net"
DRIVER_DIR='/Users/jdy62/Desktop/python_study/chromedriver_win32/chromedriver'
SAVE_DIR = 'report_1.xlsx'

#save crwaling result as excel
result=[] 
def save_excel(result):
    try:
        wb=load_workbook(SAVE_DIR)
        ws=wb['cafe']
        for idx, vi in enumerate(result):
            ws['F'+str(4+idx+1)] = vi
        wb.save(SAVE_DIR)
    
    except Exception as e:
        print(e)

#URL List extract from excel
wb=load_workbook(SAVE_DIR)
ws=wb['cafe']
get_cells=ws['G5':'G76']
report=[]
for row in get_cells:
    for cell in row:
        report.append(cell.value)
                
chrome_options=Options()
chrome_options.add_argument('disable-infobars')
chrome_options.add_argument("--user-data-dir=C:/Users/jdy62/AppData/Local/Google/Chrome/User Data/Default")
chrome_options.add_experimental_option("detach",True)
chrome_options.add_argument('--profile-directory=Profile 1')

#Login
driver = webdriver.Chrome(executable_path=DRIVER_DIR, chrome_options=chrome_options)
driver.implicitly_wait(10)
driver.get(URLN)
driver.execute_script("window.open('')")
driver.switch_to.window(driver.window_handles[1])
driver.get(URLD)

print(len(report))

#Crawling 
for i in report:
    if 'naver' in i:
        try:
            driver.switch_to.window(driver.window_handles[1])
            driver.execute_script("window.open('')")
            driver.switch_to.window(driver.window_handles[2])
            driver.get(i)
            driver.switch_to.frame('cafe_main')
            html=driver.page_source
            soup=BeautifulSoup(html,'html.parser')
            view=soup.select('span.b.m-tcol-c.reply')[1].get_text()
            print(view)
            result.append(view)
            driver.close()
            time.sleep(0.7)
        except:
            driver.switch_to_alert()
            driver.switch_to_alert().accept()
            result.append(' ')
            driver.close()
            pass
        
    elif 'daum' in i:
        try:
            driver.switch_to.window(driver.window_handles[1])
            driver.execute_script("window.open('')")
            driver.switch_to.window(driver.window_handles[2])
            driver.get(i)
            driver.switch_to.frame('down')
            html=driver.page_source
            soup=BeautifulSoup(html,'html.parser')
            view=soup.select('div.article_writer span.p11')[0].get_text()
            print(view[3:])
            result.append(view[3:])
            driver.close()
            time.sleep(0.7)
        except:
            result.append(' ')
            driver.close()            
            pass
        

    save_excel(result)
